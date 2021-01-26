import openpyxl

def readData(file, sheetName, rownum, colmno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column= colmno).value


def writeData(file, sheetName, rownum, columno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columno).value = data
    workbook.save(file)